# document_extractor.py

import os
from pathlib import Path
import logging
import re
from typing import Dict, List, Tuple, Any
import io
import tempfile
import subprocess

# 文档处理库
from docx import Document as DocxDocument
import PyPDF2
from openpyxl import load_workbook
import fitz  # PyMuPDF 作为备选PDF处理库

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ChineseDocumentProcessor:
    """处理各种格式文档的类，提取文本内容"""
    
    def __init__(self, output_dir: str = "processed_texts"):
        """初始化文档处理器
        
        Args:
            output_dir: 处理后文本的输出目录
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        # PDF处理选项
        self.pdf_retry_attempts = 2  # PDF处理失败时的重试次数
        self.robust_pdf_handling = True  # 是否启用健壮的PDF处理
    
    def process_documents(self) -> Dict[str, str]:
        """处理指定目录下的所有文档
        
        Returns:
            字典，键为文件路径，值为提取的文本内容
        """
        processed_texts = {}
        logger.info(f"Processing documents in {self.output_dir}")
        
        for root, dirs, files in os.walk(self.output_dir):
            for file in files:
                file_path = os.path.join(root, file)
                file_extension = os.path.splitext(file)[1].lower()
                
                try:
                    if file_extension == '.docx':
                        logger.info(f"Process documents: {file_path}")
                        content = self.extract_text_from_docx(file_path)
                        processed_texts[file_path] = content
                    
                    elif file_extension == '.pdf':
                        logger.info(f"Process documents: {file_path}")
                        content = self.extract_text_from_pdf(file_path)
                        processed_texts[file_path] = content
                    
                    elif file_extension in ['.xlsx', '.xls']:
                        logger.info(f"Process documents: {file_path}")
                        content = self.extract_text_from_excel(file_path)
                        processed_texts[file_path] = content
                    
                    elif file_extension == '.txt':
                        logger.info(f"Process documents: {file_path}")
                        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                            content = f.read()
                            processed_texts[file_path] = content
                except Exception as e:
                    logger.error(f"Fail to extracted {file_extension} files from: {file_path}: {e}")
        
        return processed_texts
    
    def extract_text_from_docx(self, file_path: str) -> str:
        """从DOCX文件提取文本
        
        Args:
            file_path: DOCX文件路径
            
        Returns:
            提取的文本内容
        """
        try:
            doc = DocxDocument(file_path)
            return '\n'.join([para.text for para in doc.paragraphs if para.text.strip()])
        except Exception as e:
            logger.error(f"Error extracting DOCX content from {file_path}: {e}")
            return ""
    
    def extract_text_from_pdf(self, file_path: str) -> str:
        """从PDF文件提取文本，使用多种方法尝试保证提取成功
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            提取的文本内容
        """
        text = ""
        methods_tried = []
        
        # 方法1: 使用PyPDF2
        try:
            methods_tried.append("PyPDF2")
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
                if text.strip():
                    logger.info(f"Successfully extracted text from PDF using PyPDF2: {file_path}")
                    return text
        except Exception as e:
            logger.warning(f"PyPDF2 extraction failed for {file_path}: {e}")
        
        # 如果启用了健壮处理，尝试备用方法
        if self.robust_pdf_handling:
            # 方法2: 使用PyMuPDF (fitz)
            try:
                methods_tried.append("PyMuPDF")
                doc = fitz.open(file_path)
                text = ""
                for page in doc:
                    text += page.get_text()
                if text.strip():
                    logger.info(f"Successfully extracted text from PDF using PyMuPDF: {file_path}")
                    return text
            except Exception as e:
                logger.warning(f"PyMuPDF extraction failed for {file_path}: {e}")
            
            # 方法3: 尝试修复PDF并重新提取
            try:
                methods_tried.append("PDF repair")
                repaired_pdf = self._try_repair_pdf(file_path)
                if repaired_pdf:
                    with open(repaired_pdf, 'rb') as file:
                        try:
                            reader = PyPDF2.PdfReader(file)
                            text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
                            if text.strip():
                                logger.info(f"Successfully extracted text from repaired PDF: {file_path}")
                                # 清理临时文件
                                try:
                                    os.remove(repaired_pdf)
                                except:
                                    pass
                                return text
                        except:
                            pass
            except Exception as e:
                logger.warning(f"PDF repair extraction failed for {file_path}: {e}")
            
            # 方法4: 尝试将PDF转换为文本文件并读取
            try:
                methods_tried.append("PDF to text conversion")
                text_file = self._convert_pdf_to_text(file_path)
                if text_file and os.path.exists(text_file):
                    with open(text_file, 'r', encoding='utf-8', errors='replace') as f:
                        text = f.read()
                    if text.strip():
                        logger.info(f"Successfully extracted text via conversion for PDF: {file_path}")
                        # 清理临时文件
                        try:
                            os.remove(text_file)
                        except:
                            pass
                        return text
            except Exception as e:
                logger.warning(f"PDF to text conversion failed for {file_path}: {e}")
        
        logger.error(f"All extraction methods failed for PDF {file_path}. Methods tried: {', '.join(methods_tried)}")
        return f"[PDF提取失败: {os.path.basename(file_path)}]"
    
    def _try_repair_pdf(self, file_path: str) -> str:
        """尝试修复损坏的PDF文件
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            修复后PDF的临时路径，失败则返回空字符串
        """
        # 创建临时文件用于保存修复后的PDF
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_file.close()
        repaired_path = temp_file.name
        
        try:
            # 使用PyPDF2尝试简单修复
            with open(file_path, 'rb') as file:
                try:
                    reader = PyPDF2.PdfReader(file)
                    writer = PyPDF2.PdfWriter()
                    
                    # 复制有效页面
                    for i in range(len(reader.pages)):
                        try:
                            page = reader.pages[i]
                            writer.add_page(page)
                        except:
                            logger.warning(f"Skipping corrupted page {i} in {file_path}")
                            continue
                    
                    # 写入修复后的PDF
                    with open(repaired_path, 'wb') as output_file:
                        writer.write(output_file)
                    
                    logger.info(f"PDF repair attempt successful for {file_path}")
                    return repaired_path
                except Exception as e:
                    logger.warning(f"Basic PDF repair failed: {e}")
                    
                    # 尝试读取部分PDF内容
                    file.seek(0)
                    pdf_bytes = file.read()
                    
                    # 检查PDF签名
                    if not pdf_bytes.startswith(b'%PDF-'):
                        # 尝试寻找PDF起始位置
                        start_idx = pdf_bytes.find(b'%PDF-')
                        if start_idx > 0:
                            logger.info(f"Found PDF signature at offset {start_idx}, truncating file")
                            pdf_bytes = pdf_bytes[start_idx:]
                            
                            with open(repaired_path, 'wb') as output_file:
                                output_file.write(pdf_bytes)
                            return repaired_path
            
            # 删除失败的临时文件
            if os.path.exists(repaired_path):
                os.remove(repaired_path)
            return ""
        except Exception as e:
            logger.error(f"PDF repair error: {e}")
            if os.path.exists(repaired_path):
                os.remove(repaired_path)
            return ""
    
    def _convert_pdf_to_text(self, file_path: str) -> str:
        """尝试将PDF转换为文本文件
        
        Args:
            file_path: PDF文件路径
            
        Returns:
            转换后文本文件的路径，失败则返回空字符串
        """
        # 创建临时文件用于保存转换后的文本
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
        temp_file.close()
        text_path = temp_file.name
        
        try:
            # 方法1: 使用PyMuPDF直接提取文本
            try:
                doc = fitz.open(file_path)
                text = ""
                for page in doc:
                    text += page.get_text()
                
                with open(text_path, 'w', encoding='utf-8') as f:
                    f.write(text)
                
                if os.path.getsize(text_path) > 0:
                    return text_path
            except Exception as e:
                logger.warning(f"PyMuPDF text extraction failed: {e}")
            
            # 方法2: 直接读取二进制内容并查找文本
            with open(file_path, 'rb') as file:
                content = file.read()
                # 使用正则表达式查找潜在的文本内容
                text_chunks = []
                
                # 查找PDF中的文本对象
                chunks = re.findall(br'\(([^\)]{4,})\)', content)
                for chunk in chunks:
                    try:
                        decoded = chunk.decode('utf-8', errors='ignore').strip()
                        if len(decoded) > 5:  # 忽略短内容
                            text_chunks.append(decoded)
                    except:
                        pass
                
                if text_chunks:
                    extracted_text = "\n".join(text_chunks)
                    with open(text_path, 'w', encoding='utf-8') as f:
                        f.write(extracted_text)
                    return text_path
            
            # 如果以上方法都失败，返回空字符串
            if os.path.exists(text_path):
                os.remove(text_path)
            return ""
        except Exception as e:
            logger.error(f"PDF to text conversion error: {e}")
            if os.path.exists(text_path):
                os.remove(text_path)
            return ""
    
    def extract_text_from_excel(self, file_path: str) -> str:
        """从Excel文件提取文本
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            提取的文本内容
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            text = []
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheet_text = [f"--- Sheet: {sheet} ---"]
                
                # 读取单元格数据
                for row in ws.rows:
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value))
                    if row_values:
                        sheet_text.append("\t".join(row_values))
                
                if len(sheet_text) > 1:  # 如果有内容
                    text.extend(sheet_text)
                    text.append("")  # 添加空行分隔不同工作表
            
            return '\n'.join(text)
        except Exception as e:
            logger.error(f"Error extracting Excel content from {file_path}: {e}")
            return ""

if __name__ == "__main__":
    # 测试
    processor = ChineseDocumentProcessor("./test_docs")
    results = processor.process_documents()
    print(f"Processed {len(results)} documents")
    for path, content in results.items():
        print(f"{path}: {len(content)} characters")